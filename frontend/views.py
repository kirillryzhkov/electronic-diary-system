from django.contrib.auth.views import LoginView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, ListView
from django.db.models import Avg, Count
from django.contrib import messages
from django.views.generic import CreateView, UpdateView, DeleteView
from django.core.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy

from collections import defaultdict
from django.utils.formats import date_format

from users.models import User

from grades.models import Grade
from subjects.models import Subject

from datetime import datetime

from django.http import HttpResponse
from django.views import View

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from statistics import mean
from urllib.parse import urlencode
from notifications_app.models import Notification

from academic.models import StudyGroup, Classroom, TeachingAssignment, Schedule, Attendance, Homework

from .forms import (
    GradeForm,
    SubjectForm,
    StudyGroupForm,
    ClassroomForm,
    TeachingAssignmentForm,
    ScheduleForm,
    AttendanceForm,
    HomeworkForm,
    BulkGradeAssignmentForm,
    GroupSummaryReportForm,
)




class AdminRequiredMixin(LoginRequiredMixin):
    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            raise PermissionDenied('Доступ разрешён только администратору.')
        return super().dispatch(request, *args, **kwargs)


class HomeView(TemplateView):
    template_name = 'frontend/home.html'


class CustomLoginView(LoginView):
    template_name = 'frontend/login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        return '/dashboard/'


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'frontend/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        user = self.request.user

        if user.role == 'admin':
            grades = Grade.objects.select_related('student', 'teacher', 'subject').all()
        elif user.role == 'teacher':
            grades = Grade.objects.select_related('student', 'teacher', 'subject').filter(teacher=user)
        else:
            grades = Grade.objects.select_related('student', 'teacher', 'subject').filter(student=user)

        context['grades_count'] = grades.count()
        context['subjects_count'] = Subject.objects.count()
        context['average'] = round(grades.aggregate(avg=Avg('value'))['avg'] or 0, 2)
        context['recent_grades'] = grades.order_by('-date')[:5]

        return context


class GradesView(LoginRequiredMixin, ListView):
    model = Grade
    template_name = 'frontend/grades.html'
    context_object_name = 'grades'

    def get_queryset(self):
        user = self.request.user

        queryset = Grade.objects.select_related(
            'student',
            'teacher',
            'subject',
            'student__group'
        ).order_by('-date')

        if user.role == 'teacher':
            queryset = queryset.filter(teacher=user)
        elif user.role == 'student':
            queryset = queryset.filter(student=user)

        group_id = self.request.GET.get('group')
        if group_id and user.role != 'student':
            queryset = queryset.filter(student__group_id=group_id)

        subject_id = self.request.GET.get('subject')
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)

        return queryset

    def get_groups_for_filter(self):
        user = self.request.user

        if user.role == 'teacher':
            group_ids = TeachingAssignment.objects.filter(
                teacher=user
            ).values_list('group_id', flat=True).distinct()

            return StudyGroup.objects.filter(
                id__in=group_ids
            ).order_by('name')

        if user.role == 'admin':
            return StudyGroup.objects.all().order_by('name')

        return StudyGroup.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['subjects'] = Subject.objects.order_by('name')
        context['groups'] = self.get_groups_for_filter()
        return context


class SubjectsView(LoginRequiredMixin, ListView):
    model = Subject
    template_name = 'frontend/subjects.html'
    context_object_name = 'subjects'

    def get_queryset(self):
        return Subject.objects.all()
    
class SubjectCreateFrontendView(LoginRequiredMixin, CreateView):
    model = Subject
    form_class = SubjectForm
    template_name = 'frontend/subject_form.html'
    success_url = reverse_lazy('frontend-subjects')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            raise PermissionDenied('Только администратор может создавать предметы.')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Предмет успешно добавлен.')
        return super().form_valid(form)


class SubjectUpdateFrontendView(LoginRequiredMixin, UpdateView):
    model = Subject
    form_class = SubjectForm
    template_name = 'frontend/subject_form.html'
    success_url = reverse_lazy('frontend-subjects')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            raise PermissionDenied('Только администратор может редактировать предметы.')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Предмет успешно обновлён.')
        return super().form_valid(form)


class SubjectDeleteFrontendView(LoginRequiredMixin, DeleteView):
    model = Subject
    template_name = 'frontend/subject_confirm_delete.html'
    success_url = reverse_lazy('frontend-subjects')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            raise PermissionDenied('Только администратор может удалять предметы.')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Предмет успешно удалён.')
        return super().form_valid(form)


class StatsView(LoginRequiredMixin, TemplateView):
    template_name = 'frontend/stats.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'teacher':
            from django.shortcuts import redirect
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        subjects = Subject.objects.annotate(
            total_grades=Count('grades'),
            average=Avg('grades__value')
        )

        context['subjects'] = subjects
        return context
    
class GradeCreateFrontendView(LoginRequiredMixin, CreateView):
    model = Grade
    template_name = 'frontend/grade_form.html'
    form_class = GradeForm
    success_url = reverse_lazy('frontend-grades')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'teacher':
            raise PermissionDenied('Только учитель может создавать оценки.')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.teacher = self.request.user
        response = super().form_valid(form)

        Notification.objects.create(
            user=self.object.student,
            notification_type='grade',
            title='Новая оценка',
            message=(
                f'Вам поставили оценку {self.object.value} '
                f'по предмету "{self.object.subject.name}".'
            )
        )

        messages.success(self.request, 'Оценка успешно добавлена.')
        return response


class GradeUpdateFrontendView(LoginRequiredMixin, UpdateView):
    model = Grade
    template_name = 'frontend/grade_form.html'
    form_class = GradeForm
    success_url = reverse_lazy('frontend-grades')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'teacher':
            raise PermissionDenied('Только учитель может редактировать оценки.')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return Grade.objects.select_related(
            'student',
            'teacher',
            'subject'
        ).filter(
            teacher=self.request.user
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, 'Оценка успешно обновлена.')
        return super().form_valid(form)


class GradeDeleteFrontendView(LoginRequiredMixin, DeleteView):
    model = Grade
    template_name = 'frontend/grade_confirm_delete.html'
    success_url = reverse_lazy('frontend-grades')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'teacher':
            raise PermissionDenied('Только учитель может удалять оценки.')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Оценка успешно удалена.')
        return super().form_valid(form)
    
    def get_queryset(self):
        user = self.request.user

        queryset = Grade.objects.select_related('student', 'teacher', 'subject')

        if user.role == 'teacher':
            queryset = queryset.filter(teacher=user)

        return queryset
    

class StudyGroupListView(LoginRequiredMixin, ListView):
    model = StudyGroup
    template_name = 'frontend/groups.html'
    context_object_name = 'groups'

    def get_queryset(self):
        return StudyGroup.objects.select_related('curator').all()


class StudyGroupCreateView(AdminRequiredMixin, CreateView):
    model = StudyGroup
    form_class = StudyGroupForm
    template_name = 'frontend/group_form.html'
    success_url = reverse_lazy('frontend-groups')

    def form_valid(self, form):
        messages.success(self.request, 'Группа успешно добавлена.')
        return super().form_valid(form)


class StudyGroupUpdateView(AdminRequiredMixin, UpdateView):
    model = StudyGroup
    form_class = StudyGroupForm
    template_name = 'frontend/group_form.html'
    success_url = reverse_lazy('frontend-groups')

    def form_valid(self, form):
        messages.success(self.request, 'Группа успешно обновлена.')
        return super().form_valid(form)


class StudyGroupDeleteView(AdminRequiredMixin, DeleteView):
    model = StudyGroup
    template_name = 'frontend/group_confirm_delete.html'
    success_url = reverse_lazy('frontend-groups')

    def form_valid(self, form):
        messages.success(self.request, 'Группа успешно удалена.')
        return super().form_valid(form)


class ClassroomListView(LoginRequiredMixin, ListView):
    model = Classroom
    template_name = 'frontend/classrooms.html'
    context_object_name = 'classrooms'

    def get_queryset(self):
        return Classroom.objects.all()


class ClassroomCreateView(AdminRequiredMixin, CreateView):
    model = Classroom
    form_class = ClassroomForm
    template_name = 'frontend/classroom_form.html'
    success_url = reverse_lazy('frontend-classrooms')

    def form_valid(self, form):
        messages.success(self.request, 'Кабинет успешно добавлен.')
        return super().form_valid(form)


class ClassroomUpdateView(AdminRequiredMixin, UpdateView):
    model = Classroom
    form_class = ClassroomForm
    template_name = 'frontend/classroom_form.html'
    success_url = reverse_lazy('frontend-classrooms')

    def form_valid(self, form):
        messages.success(self.request, 'Кабинет успешно обновлён.')
        return super().form_valid(form)


class ClassroomDeleteView(AdminRequiredMixin, DeleteView):
    model = Classroom
    template_name = 'frontend/classroom_confirm_delete.html'
    success_url = reverse_lazy('frontend-classrooms')

    def form_valid(self, form):
        messages.success(self.request, 'Кабинет успешно удалён.')
        return super().form_valid(form)


class TeachingAssignmentListView(LoginRequiredMixin, ListView):
    model = TeachingAssignment
    template_name = 'frontend/assignments.html'
    context_object_name = 'assignments'

    def get_queryset(self):
        user = self.request.user

        queryset = TeachingAssignment.objects.select_related(
            'teacher',
            'subject',
            'group',
            'classroom'
        )

        if user.role == 'teacher':
            queryset = queryset.filter(teacher=user)

        if user.role == 'student' and user.group:
            queryset = queryset.filter(group=user.group)

        return queryset


class TeachingAssignmentCreateView(AdminRequiredMixin, CreateView):
    model = TeachingAssignment
    form_class = TeachingAssignmentForm
    template_name = 'frontend/assignment_form.html'
    success_url = reverse_lazy('frontend-assignments')

    def form_valid(self, form):
        messages.success(self.request, 'Назначение преподавателя успешно добавлено.')
        return super().form_valid(form)


class TeachingAssignmentUpdateView(AdminRequiredMixin, UpdateView):
    model = TeachingAssignment
    form_class = TeachingAssignmentForm
    template_name = 'frontend/assignment_form.html'
    success_url = reverse_lazy('frontend-assignments')

    def form_valid(self, form):
        messages.success(self.request, 'Назначение преподавателя успешно обновлено.')
        return super().form_valid(form)


class TeachingAssignmentDeleteView(AdminRequiredMixin, DeleteView):
    model = TeachingAssignment
    template_name = 'frontend/assignment_confirm_delete.html'
    success_url = reverse_lazy('frontend-assignments')

    def form_valid(self, form):
        messages.success(self.request, 'Назначение преподавателя успешно удалено.')
        return super().form_valid(form)
    
class TeacherGroupsView(LoginRequiredMixin, TemplateView):
    template_name = 'frontend/teacher_groups.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'teacher':
            raise PermissionDenied('Эта страница доступна только преподавателю.')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        teacher = self.request.user

        groups = StudyGroup.objects.filter(
            teaching_assignments__teacher=teacher
        ).distinct().order_by('name')

        groups_data = []

        for group in groups:
            assignments = TeachingAssignment.objects.select_related(
                'subject',
                'classroom'
            ).filter(
                teacher=teacher,
                group=group
            )

            students = User.objects.filter(
                role='student',
                group=group
            ).order_by('username')

            grades = Grade.objects.filter(
                teacher=teacher,
                student__group=group
            )

            groups_data.append({
                'group': group,
                'assignments': assignments,
                'students_count': students.count(),
                'grades_count': grades.count(),
                'average': round(grades.aggregate(avg=Avg('value'))['avg'] or 0, 2),
            })

        context['groups_data'] = groups_data
        return context


class TeacherGroupDetailView(LoginRequiredMixin, TemplateView):
    template_name = 'frontend/teacher_group_detail.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'teacher':
            raise PermissionDenied('Эта страница доступна только преподавателю.')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        teacher = self.request.user
        group = get_object_or_404(StudyGroup, id=self.kwargs['group_id'])

        has_access = TeachingAssignment.objects.filter(
            teacher=teacher,
            group=group
        ).exists()

        if not has_access:
            raise PermissionDenied('Вы не ведёте эту группу.')

        assignments = TeachingAssignment.objects.select_related(
            'subject',
            'classroom'
        ).filter(
            teacher=teacher,
            group=group
        )

        students = User.objects.filter(
            role='student',
            group=group
        ).order_by('username')

        grades = Grade.objects.select_related(
            'student',
            'teacher',
            'subject'
        ).filter(
            teacher=teacher,
            student__group=group
        ).order_by('-date')

        context['group'] = group
        context['assignments'] = assignments
        context['students'] = students
        context['grades'] = grades
        context['grades_count'] = grades.count()
        context['students_count'] = students.count()
        context['average'] = round(grades.aggregate(avg=Avg('value'))['avg'] or 0, 2)

        return context
    
class ProfileView(LoginRequiredMixin, TemplateView):
    template_name = 'frontend/profile.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        user = self.request.user

        context['profile_user'] = user

        if user.role == 'student':
            grades = Grade.objects.select_related(
                'student',
                'teacher',
                'subject'
            ).filter(student=user).order_by('-date')

            context['grades_count'] = grades.count()
            context['average'] = round(grades.aggregate(avg=Avg('value'))['avg'] or 0, 2)
            context['recent_grades'] = grades[:5]

        elif user.role == 'teacher':
            grades = Grade.objects.select_related(
                'student',
                'teacher',
                'subject'
            ).filter(teacher=user).order_by('-date')

            assignments = TeachingAssignment.objects.select_related(
                'subject',
                'group',
                'classroom'
            ).filter(teacher=user)

            groups = StudyGroup.objects.filter(
                teaching_assignments__teacher=user
            ).distinct()

            context['grades_count'] = grades.count()
            context['average'] = round(grades.aggregate(avg=Avg('value'))['avg'] or 0, 2)
            context['recent_grades'] = grades[:5]
            context['assignments'] = assignments
            context['groups_count'] = groups.count()

        elif user.role == 'admin':
            context['users_count'] = User.objects.count()
            context['students_count'] = User.objects.filter(role='student').count()
            context['teachers_count'] = User.objects.filter(role='teacher').count()
            context['subjects_count'] = Subject.objects.count()
            context['groups_count'] = StudyGroup.objects.count()
            context['classrooms_count'] = Classroom.objects.count()
            context['assignments_count'] = TeachingAssignment.objects.count()
            context['grades_count'] = Grade.objects.count()

        return context
    
class ScheduleListView(LoginRequiredMixin, ListView):
    model = Schedule
    template_name = 'frontend/schedule.html'
    context_object_name = 'schedule_items'

    def get_queryset(self):
        user = self.request.user

        queryset = Schedule.objects.select_related(
            'assignment',
            'assignment__teacher',
            'assignment__subject',
            'assignment__group',
            'assignment__classroom',
        ).order_by('day', 'start_time')

        if user.role == 'teacher':
            queryset = queryset.filter(assignment__teacher=user)

        elif user.role == 'student':
            if user.group:
                queryset = queryset.filter(assignment__group=user.group)
            else:
                queryset = queryset.none()

        return queryset


class ScheduleCreateView(AdminRequiredMixin, CreateView):
    model = Schedule
    form_class = ScheduleForm
    template_name = 'frontend/schedule_form.html'
    success_url = reverse_lazy('frontend-schedule')

    def form_valid(self, form):
        messages.success(self.request, 'Занятие успешно добавлено в расписание.')
        return super().form_valid(form)


class ScheduleUpdateView(AdminRequiredMixin, UpdateView):
    model = Schedule
    form_class = ScheduleForm
    template_name = 'frontend/schedule_form.html'
    success_url = reverse_lazy('frontend-schedule')

    def form_valid(self, form):
        messages.success(self.request, 'Занятие в расписании успешно обновлено.')
        return super().form_valid(form)


class ScheduleDeleteView(AdminRequiredMixin, DeleteView):
    model = Schedule
    template_name = 'frontend/schedule_confirm_delete.html'
    success_url = reverse_lazy('frontend-schedule')

    def form_valid(self, form):
        messages.success(self.request, 'Занятие удалено из расписания.')
        return super().form_valid(form)
    
class AttendanceListView(LoginRequiredMixin, ListView):
    model = Attendance
    template_name = 'frontend/attendance.html'
    context_object_name = 'attendance_records'

    def get_queryset(self):
        user = self.request.user

        queryset = Attendance.objects.select_related(
            'student',
            'assignment',
            'assignment__teacher',
            'assignment__subject',
            'assignment__group',
            'assignment__classroom',
        ).order_by('-date')

        if user.role == 'teacher':
            queryset = queryset.filter(assignment__teacher=user)
        elif user.role == 'student':
            queryset = queryset.filter(student=user)

        group_id = self.request.GET.get('group')
        if group_id and user.role != 'student':
            queryset = queryset.filter(assignment__group_id=group_id)

        subject_id = self.request.GET.get('subject')
        if subject_id:
            queryset = queryset.filter(assignment__subject_id=subject_id)

        status_value = self.request.GET.get('status')
        if status_value:
            queryset = queryset.filter(status=status_value)

        return queryset

    def get_groups_for_filter(self):
        user = self.request.user

        if user.role == 'teacher':
            group_ids = TeachingAssignment.objects.filter(
                teacher=user
            ).values_list('group_id', flat=True).distinct()

            return StudyGroup.objects.filter(
                id__in=group_ids
            ).order_by('name')

        if user.role == 'admin':
            return StudyGroup.objects.all().order_by('name')

        return StudyGroup.objects.none()

    def get_subjects_for_filter(self):
        user = self.request.user

        if user.role == 'teacher':
            subject_ids = TeachingAssignment.objects.filter(
                teacher=user
            ).values_list('subject_id', flat=True).distinct()

            return Subject.objects.filter(
                id__in=subject_ids
            ).order_by('name')

        if user.role == 'admin':
            return Subject.objects.all().order_by('name')

        if user.role == 'student' and user.group:
            subject_ids = TeachingAssignment.objects.filter(
                group=user.group
            ).values_list('subject_id', flat=True).distinct()

            return Subject.objects.filter(
                id__in=subject_ids
            ).order_by('name')

        return Subject.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['groups'] = self.get_groups_for_filter()
        context['subjects'] = self.get_subjects_for_filter()
        context['status_choices'] = Attendance.STATUS_CHOICES
        return context


class AttendanceCreateView(LoginRequiredMixin, CreateView):
    model = Attendance
    form_class = AttendanceForm
    template_name = 'frontend/attendance_form.html'
    success_url = reverse_lazy('frontend-attendance')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ['admin', 'teacher']:
            raise PermissionDenied('Только администратор или преподаватель может отмечать посещаемость.')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, 'Посещаемость успешно добавлена.')
        return super().form_valid(form)


class AttendanceUpdateView(LoginRequiredMixin, UpdateView):
    model = Attendance
    form_class = AttendanceForm
    template_name = 'frontend/attendance_form.html'
    success_url = reverse_lazy('frontend-attendance')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ['admin', 'teacher']:
            raise PermissionDenied('Только администратор или преподаватель может редактировать посещаемость.')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user

        queryset = Attendance.objects.select_related(
            'student',
            'assignment',
            'assignment__teacher',
            'assignment__subject',
            'assignment__group',
        )

        if user.role == 'teacher':
            queryset = queryset.filter(assignment__teacher=user)

        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, 'Посещаемость успешно обновлена.')
        return super().form_valid(form)


class AttendanceDeleteView(LoginRequiredMixin, DeleteView):
    model = Attendance
    template_name = 'frontend/attendance_confirm_delete.html'
    success_url = reverse_lazy('frontend-attendance')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ['admin', 'teacher']:
            raise PermissionDenied('Только администратор или преподаватель может удалять посещаемость.')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user

        queryset = Attendance.objects.select_related(
            'student',
            'assignment',
            'assignment__teacher',
            'assignment__subject',
            'assignment__group',
        )

        if user.role == 'teacher':
            queryset = queryset.filter(assignment__teacher=user)

        return queryset

    def form_valid(self, form):
        messages.success(self.request, 'Запись посещаемости удалена.')
        return super().form_valid(form)
    
class HomeworkListView(LoginRequiredMixin, ListView):
    model = Homework
    template_name = 'frontend/homework.html'
    context_object_name = 'homeworks'

    def get_queryset(self):
        user = self.request.user

        queryset = Homework.objects.select_related(
            'assignment',
            'assignment__teacher',
            'assignment__subject',
            'assignment__group',
            'assignment__classroom',
        ).order_by('-created_at')

        if user.role == 'teacher':
            queryset = queryset.filter(assignment__teacher=user)

        elif user.role == 'student':
            if user.group:
                queryset = queryset.filter(assignment__group=user.group)
            else:
                queryset = queryset.none()

        return queryset


class HomeworkCreateView(LoginRequiredMixin, CreateView):
    model = Homework
    form_class = HomeworkForm
    template_name = 'frontend/homework_form.html'
    success_url = reverse_lazy('frontend-homework')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ['admin', 'teacher']:
            raise PermissionDenied('Только администратор или преподаватель может добавлять домашние задания.')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)

        students = User.objects.filter(
            role='student',
            group=self.object.assignment.group
        )

        notifications = [
            Notification(
                user=student,
                notification_type='homework',
                title='Новое домашнее задание',
                message=(
                    f'Добавлено новое домашнее задание по предмету '
                    f'"{self.object.assignment.subject.name}": {self.object.title}'
                )
            )
            for student in students
        ]

        Notification.objects.bulk_create(notifications)

        messages.success(self.request, 'Домашнее задание успешно добавлено.')
        return response


class HomeworkUpdateView(LoginRequiredMixin, UpdateView):
    model = Homework
    form_class = HomeworkForm
    template_name = 'frontend/homework_form.html'
    success_url = reverse_lazy('frontend-homework')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ['admin', 'teacher']:
            raise PermissionDenied('Только администратор или преподаватель может редактировать домашние задания.')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user

        queryset = Homework.objects.select_related(
            'assignment',
            'assignment__teacher',
            'assignment__subject',
            'assignment__group',
            'assignment__classroom',
        )

        if user.role == 'teacher':
            queryset = queryset.filter(assignment__teacher=user)

        return queryset

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, 'Домашнее задание успешно обновлено.')
        return super().form_valid(form)


class HomeworkDeleteView(LoginRequiredMixin, DeleteView):
    model = Homework
    template_name = 'frontend/homework_confirm_delete.html'
    success_url = reverse_lazy('frontend-homework')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ['admin', 'teacher']:
            raise PermissionDenied('Только администратор или преподаватель может удалять домашние задания.')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user

        queryset = Homework.objects.select_related(
            'assignment',
            'assignment__teacher',
            'assignment__subject',
            'assignment__group',
            'assignment__classroom',
        )

        if user.role == 'teacher':
            queryset = queryset.filter(assignment__teacher=user)

        return queryset

    def form_valid(self, form):
        messages.success(self.request, 'Домашнее задание удалено.')
        return super().form_valid(form)
    

class GroupGradeEntryView(LoginRequiredMixin, TemplateView):
    template_name = 'frontend/group_grade_entry.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'teacher':
            raise PermissionDenied('Только преподаватель может выставлять оценки группе.')
        return super().dispatch(request, *args, **kwargs)

    def get_selected_assignment(self):
        assignment_id = self.request.GET.get('assignment') or self.request.POST.get('assignment')

        if not assignment_id:
            return None

        return get_object_or_404(
            TeachingAssignment.objects.select_related(
                'group',
                'subject',
                'classroom',
                'teacher'
            ),
            pk=assignment_id,
            teacher=self.request.user
        )

    def get_students(self, assignment):
        if not assignment:
            return User.objects.none()

        return User.objects.filter(
            role='student',
            group=assignment.group
        ).order_by('last_name', 'first_name', 'username')

    def get_last_grades_map(self, assignment, students):
        if not assignment:
            return {}

        student_ids = list(students.values_list('id', flat=True))

        last_grades = Grade.objects.select_related('subject').filter(
            student_id__in=student_ids,
            subject=assignment.subject
        ).order_by('student_id', '-date')

        result = {}
        for grade in last_grades:
            if grade.student_id not in result:
                result[grade.student_id] = grade

        return result

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        selected_assignment = self.get_selected_assignment()
        students = self.get_students(selected_assignment)

        form_data = self.request.GET if self.request.GET.get('assignment') else None
        form = BulkGradeAssignmentForm(form_data, user=self.request.user)

        if selected_assignment and not form.is_bound:
            form = BulkGradeAssignmentForm(
                initial={
                    'assignment': selected_assignment,
                    'grade_type': 'current',
                },
                user=self.request.user
            )

        context['form'] = form
        context['selected_assignment'] = selected_assignment
        context['students'] = students
        context['last_grades_map'] = self.get_last_grades_map(selected_assignment, students) or {}

        return context

    def post(self, request, *args, **kwargs):
        selected_assignment = self.get_selected_assignment()
        form = BulkGradeAssignmentForm(request.POST, user=request.user)

        if not form.is_valid():
            context = self.get_context_data()
            context['form'] = form
            return self.render_to_response(context)

        if not selected_assignment:
            messages.error(request, 'Сначала выберите группу и предмет.')
            return redirect('frontend-group-grade-entry')

        students = self.get_students(selected_assignment)

        grade_type = form.cleaned_data['grade_type']
        month = form.cleaned_data['month'] or None
        semester = form.cleaned_data['semester'] or None

        if month:
            month = int(month)

        if semester:
            semester = int(semester)

        created_count = 0
        skipped_count = 0
        invalid_students = []

        for student in students:
            value = request.POST.get(f'value_{student.id}', '').strip()
            comment = request.POST.get(f'comment_{student.id}', '').strip()

            if not value:
                skipped_count += 1
                continue

            try:
                value_int = int(value)
            except ValueError:
                invalid_students.append(student.full_name)
                continue

            if value_int < 1 or value_int > 5:
                invalid_students.append(student.full_name)
                continue

            grade = Grade.objects.create(
                student=student,
                teacher=request.user,
                subject=selected_assignment.subject,
                value=value_int,
                comment=comment,
                grade_type=grade_type,
                month=month,
                semester=semester
            )

            Notification.objects.create(
                user=student,
                notification_type='grade',
                title='Новая оценка',
                message=(
                    f'Вам поставили оценку {grade.value} '
                    f'по предмету "{grade.subject.name}".'
                )
            )
            created_count += 1

        if created_count:
            messages.success(
                request,
                f'Успешно добавлено оценок: {created_count}.'
            )

        if skipped_count and created_count == 0 and not invalid_students:
            messages.warning(
                request,
                'Вы не заполнили ни одной оценки.'
            )

        if invalid_students:
            names = ', '.join(invalid_students[:5])
            messages.error(
                request,
                f'Некорректные оценки у студентов: {names}.'
            )

        redirect_url = (
            f"{reverse('frontend-group-grade-entry')}"
            f"?assignment={selected_assignment.id}"
            f"&grade_type={grade_type}"
            f"&month={month or ''}"
            f"&semester={semester or ''}"
        )
        return redirect(redirect_url)
    
class GradeJournalView(LoginRequiredMixin, TemplateView):
    template_name = 'frontend/grade_journal.html'

    def get_base_queryset(self):
        user = self.request.user

        queryset = Grade.objects.select_related(
            'student',
            'teacher',
            'subject',
            'student__group'
        ).order_by('-date')

        if user.role == 'teacher':
            queryset = queryset.filter(teacher=user)
        elif user.role == 'student':
            queryset = queryset.filter(student=user)

        return queryset

    def get_queryset(self):
        queryset = self.get_base_queryset()

        group_id = self.request.GET.get('group')
        if group_id and self.request.user.role != 'student':
            queryset = queryset.filter(student__group_id=group_id)

        subject_id = self.request.GET.get('subject')
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)

        grade_type = self.request.GET.get('grade_type')
        if grade_type:
            queryset = queryset.filter(grade_type=grade_type)

        student_id = self.request.GET.get('student')
        if student_id and self.request.user.role != 'student':
            queryset = queryset.filter(student_id=student_id)

        return queryset

    def get_groups_for_filter(self):
        user = self.request.user

        if user.role == 'teacher':
            group_ids = TeachingAssignment.objects.filter(
                teacher=user
            ).values_list('group_id', flat=True).distinct()

            return StudyGroup.objects.filter(
                id__in=group_ids
            ).order_by('name')

        if user.role == 'admin':
            return StudyGroup.objects.all().order_by('name')

        return StudyGroup.objects.none()

    def get_students_for_filter(self):
        user = self.request.user
        selected_group_id = self.request.GET.get('group')

        queryset = User.objects.none()

        if user.role == 'teacher':
            group_ids = TeachingAssignment.objects.filter(
                teacher=user
            ).values_list('group_id', flat=True).distinct()

            queryset = User.objects.filter(
                role='student',
                group_id__in=group_ids
            )

        elif user.role == 'admin':
            queryset = User.objects.filter(role='student')

        if selected_group_id and user.role != 'student':
            queryset = queryset.filter(group_id=selected_group_id)

        return queryset.order_by('last_name', 'first_name', 'username')

    def build_grade_type_filter_url(self, grade_type=None):
        params = {}

        group_id = self.request.GET.get('group')
        subject_id = self.request.GET.get('subject')
        student_id = self.request.GET.get('student')

        if group_id and self.request.user.role != 'student':
            params['group'] = group_id

        if subject_id:
            params['subject'] = subject_id

        if student_id and self.request.user.role != 'student':
            params['student'] = student_id

        if grade_type:
            params['grade_type'] = grade_type

        query_string = urlencode(params)
        base_url = reverse('frontend-grade-journal')

        if query_string:
            return f'{base_url}?{query_string}'
        return base_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        grades = self.get_queryset()
        grouped = defaultdict(list)

        for grade in grades:
            grouped[grade.date.date()].append(grade)

        grouped_days = []
        for day in sorted(grouped.keys(), reverse=True):
            grouped_days.append({
                'date': day,
                'weekday': date_format(day, 'l'),
                'grades': grouped[day],
                'count': len(grouped[day]),
            })

        selected_group_id = self.request.GET.get('group')
        selected_subject_id = self.request.GET.get('subject')
        selected_grade_type = self.request.GET.get('grade_type')
        selected_student_id = self.request.GET.get('student')

        selected_group = None
        if selected_group_id and self.request.user.role != 'student':
            selected_group = StudyGroup.objects.filter(id=selected_group_id).first()

        selected_subject = None
        if selected_subject_id:
            selected_subject = Subject.objects.filter(id=selected_subject_id).first()

        selected_student = None
        if selected_student_id and self.request.user.role != 'student':
            selected_student = User.objects.filter(id=selected_student_id).first()

        grade_type_map = dict(Grade.GRADE_TYPE_CHOICES)

        context['grouped_days'] = grouped_days
        context['subjects'] = Subject.objects.order_by('name')
        context['groups'] = self.get_groups_for_filter()
        context['students'] = self.get_students_for_filter()
        context['grade_types'] = Grade.GRADE_TYPE_CHOICES

        context['total_grades'] = grades.count()
        context['total_days'] = len(grouped_days)
        context['average_grade'] = round(grades.aggregate(avg=Avg('value'))['avg'] or 0, 2) if grades.exists() else '—'

        context['current_count'] = grades.filter(grade_type='current').count()
        context['monthly_count'] = grades.filter(grade_type='monthly').count()
        context['semester_count'] = grades.filter(grade_type='semester').count()
        context['exam_count'] = grades.filter(grade_type='exam').count()

        context['selected_group'] = selected_group
        context['selected_subject'] = selected_subject
        context['selected_student'] = selected_student
        context['selected_grade_type'] = selected_grade_type
        context['selected_grade_type_label'] = grade_type_map.get(selected_grade_type)

        context['type_filter_urls'] = {
            'all': self.build_grade_type_filter_url(),
            'current': self.build_grade_type_filter_url('current'),
            'monthly': self.build_grade_type_filter_url('monthly'),
            'semester': self.build_grade_type_filter_url('semester'),
            'exam': self.build_grade_type_filter_url('exam'),
        }

        return context
    

class GradeExportExcelView(LoginRequiredMixin, View):
    WEEKDAYS_RU = {
        0: 'Понедельник',
        1: 'Вторник',
        2: 'Среда',
        3: 'Четверг',
        4: 'Пятница',
        5: 'Суббота',
        6: 'Воскресенье',
    }

    def get_queryset(self):
        user = self.request.user

        queryset = Grade.objects.select_related(
            'student',
            'teacher',
            'subject'
        ).order_by('-date')

        if user.role == 'teacher':
            queryset = queryset.filter(teacher=user)
        elif user.role == 'student':
            queryset = queryset.filter(student=user)

        subject_id = self.request.GET.get('subject')
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)

        grade_type = self.request.GET.get('grade_type')
        if grade_type:
            queryset = queryset.filter(grade_type=grade_type)

        return queryset

    def get(self, request, *args, **kwargs):
        grades = self.get_queryset()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Grades'

        title_fill = PatternFill(fill_type='solid', fgColor='1E3A8A')
        header_fill = PatternFill(fill_type='solid', fgColor='DBEAFE')

        thin_side = Side(style='thin', color='D1D5DB')
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        title_font = Font(color='FFFFFF', bold=True, size=14)
        header_font = Font(bold=True, size=11)
        body_font = Font(size=10)

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        if request.user.role == 'student':
            export_title = 'Экспорт моих оценок'
        elif request.user.role == 'teacher':
            export_title = 'Экспорт оценок преподавателя'
        else:
            export_title = 'Экспорт всех оценок'

        headers = [
            'Дата',
            'День недели',
            'Студент',
            'Преподаватель',
            'Предмет',
            'Тип оценки',
            'Период',
            'Оценка',
            'Комментарий',
        ]

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = sheet.cell(row=1, column=1, value=export_title)
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        title_cell.border = border

        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        row_num = 3

        for grade in grades:
            weekday_name = self.WEEKDAYS_RU.get(grade.date.weekday(), '—')

            row_data = [
                grade.date.strftime('%d.%m.%Y'),
                weekday_name,
                grade.student.full_name,
                grade.teacher.full_name if grade.teacher else '—',
                grade.subject.name,
                grade.get_grade_type_display(),
                grade.period_label,
                grade.value,
                grade.comment or '—',
            ]

            for col_num, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.font = body_font
                cell.border = border

                if col_num in [1, 2, 8]:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment

            row_num += 1

        column_widths = {
            1: 14,
            2: 18,
            3: 28,
            4: 28,
            5: 22,
            6: 22,
            7: 18,
            8: 10,
            9: 45,
        }

        for col_num, width in column_widths.items():
            sheet.column_dimensions[get_column_letter(col_num)].width = width

        sheet.freeze_panes = 'A3'
        sheet.auto_filter.ref = f'A2:I{max(row_num - 1, 2)}'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f'grades_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response
    
class GradeSummaryBaseMixin(LoginRequiredMixin):
    MONTHS_RU = {
        1: 'Январь',
        2: 'Февраль',
        3: 'Март',
        4: 'Апрель',
        5: 'Май',
        6: 'Июнь',
        7: 'Июль',
        8: 'Август',
        9: 'Сентябрь',
        10: 'Октябрь',
        11: 'Ноябрь',
        12: 'Декабрь',
    }

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ['admin', 'teacher']:
            raise PermissionDenied('Итоговый отчёт доступен только администратору и преподавателю.')
        return super().dispatch(request, *args, **kwargs)

    def get_allowed_assignments(self):
        queryset = TeachingAssignment.objects.select_related(
            'teacher',
            'subject',
            'group',
            'classroom'
        )

        if self.request.user.role == 'teacher':
            queryset = queryset.filter(teacher=self.request.user)

        return queryset

    def get_selected_assignment(self):
        assignment_id = self.request.GET.get('assignment')

        if not assignment_id:
            return None

        return get_object_or_404(
            self.get_allowed_assignments(),
            pk=assignment_id
        )

    def get_students(self, assignment):
        if not assignment:
            return User.objects.none()

        return User.objects.filter(
            role='student',
            group=assignment.group
        ).order_by('last_name', 'first_name', 'username')

    def get_grades(self, assignment):
        if not assignment:
            return Grade.objects.none()

        return Grade.objects.select_related(
            'student',
            'teacher',
            'subject'
        ).filter(
            student__group=assignment.group,
            subject=assignment.subject,
            teacher=assignment.teacher
        ).order_by('student__last_name', 'student__first_name', '-date')

    def get_latest_grade_value(self, grades, grade_type, month=None, semester=None):
        filtered = []

        for grade in grades:
            if grade.grade_type != grade_type:
                continue
            if month is not None and grade.month != month:
                continue
            if semester is not None and grade.semester != semester:
                continue
            filtered.append(grade)

        if not filtered:
            return '—'

        return filtered[0].value

    def build_report_data(self, assignment):
        students = list(self.get_students(assignment))
        grades = list(self.get_grades(assignment))

        monthly_numbers = sorted({
            grade.month for grade in grades
            if grade.grade_type == 'monthly' and grade.month
        })

        grades_by_student = defaultdict(list)
        for grade in grades:
            grades_by_student[grade.student_id].append(grade)

        rows = []

        for index, student in enumerate(students, start=1):
            student_grades = grades_by_student.get(student.id, [])

            current_values = [
                grade.value for grade in student_grades
                if grade.grade_type == 'current'
            ]

            current_average = round(mean(current_values), 2) if current_values else '—'
            overall_average = round(mean([grade.value for grade in student_grades]), 2) if student_grades else '—'

            monthly_values = {
                month: self.get_latest_grade_value(student_grades, 'monthly', month=month)
                for month in monthly_numbers
            }

            row = {
                'index': index,
                'student': student,
                'current_average': current_average,
                'monthly_values': monthly_values,
                'semester_1': self.get_latest_grade_value(student_grades, 'semester', semester=1),
                'exam_1': self.get_latest_grade_value(student_grades, 'exam', semester=1),
                'semester_2': self.get_latest_grade_value(student_grades, 'semester', semester=2),
                'exam_2': self.get_latest_grade_value(student_grades, 'exam', semester=2),
                'overall_average': overall_average,
                'grades_count': len(student_grades),
            }
            rows.append(row)

        return {
            'students': students,
            'grades': grades,
            'monthly_numbers': monthly_numbers,
            'monthly_labels': [self.MONTHS_RU.get(month, str(month)) for month in monthly_numbers],
            'rows': rows,
        }


class GradeSummaryReportView(GradeSummaryBaseMixin, TemplateView):
    template_name = 'frontend/grade_summary_report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        selected_assignment = self.get_selected_assignment()
        form = GroupSummaryReportForm(
            self.request.GET or None,
            user=self.request.user
        )

        context['form'] = form
        context['selected_assignment'] = selected_assignment

        if selected_assignment:
            report_data = self.build_report_data(selected_assignment)
            context.update(report_data)

        return context


class GradeSummaryExcelExportView(GradeSummaryBaseMixin, View):
    def apply_grade_style(self, cell, value, border, center_alignment, body_font):
        cell.font = body_font
        cell.border = border
        cell.alignment = center_alignment

        if value in ['—', '', None]:
            return

        try:
            numeric_value = float(str(value).replace(',', '.'))
        except ValueError:
            return

        if numeric_value >= 5:
            cell.fill = PatternFill(fill_type='solid', fgColor='DCFCE7')
            cell.font = Font(size=10, bold=True, color='166534')
        elif numeric_value >= 4:
            cell.fill = PatternFill(fill_type='solid', fgColor='DBEAFE')
            cell.font = Font(size=10, bold=True, color='1E40AF')
        elif numeric_value >= 3:
            cell.fill = PatternFill(fill_type='solid', fgColor='FEF3C7')
            cell.font = Font(size=10, bold=True, color='92400E')
        else:
            cell.fill = PatternFill(fill_type='solid', fgColor='FEE2E2')
            cell.font = Font(size=10, bold=True, color='991B1B')

    def apply_average_style(self, cell, value, border, center_alignment, body_font):
        cell.font = body_font
        cell.border = border
        cell.alignment = center_alignment

        if value in ['—', '', None]:
            return

        try:
            numeric_value = float(str(value).replace(',', '.'))
        except ValueError:
            return

        if numeric_value >= 4.5:
            cell.fill = PatternFill(fill_type='solid', fgColor='BBF7D0')
            cell.font = Font(size=10, bold=True, color='166534')
        elif numeric_value >= 4:
            cell.fill = PatternFill(fill_type='solid', fgColor='BFDBFE')
            cell.font = Font(size=10, bold=True, color='1E40AF')
        elif numeric_value >= 3:
            cell.fill = PatternFill(fill_type='solid', fgColor='FDE68A')
            cell.font = Font(size=10, bold=True, color='92400E')
        else:
            cell.fill = PatternFill(fill_type='solid', fgColor='FECACA')
            cell.font = Font(size=10, bold=True, color='991B1B')

    def get(self, request, *args, **kwargs):
        assignment = self.get_selected_assignment()

        if not assignment:
            messages.error(request, 'Сначала выберите группу и предмет.')
            return redirect('frontend-grade-summary-report')

        report_data = self.build_report_data(assignment)
        workbook = Workbook()

        title_fill = PatternFill(fill_type='solid', fgColor='1E3A8A')
        header_fill = PatternFill(fill_type='solid', fgColor='DBEAFE')
        section_fill = PatternFill(fill_type='solid', fgColor='E0E7FF')

        thin_side = Side(style='thin', color='D1D5DB')
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        title_font = Font(color='FFFFFF', bold=True, size=14)
        header_font = Font(bold=True, size=11)
        body_font = Font(size=10)

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        summary_sheet = workbook.active
        summary_sheet.title = 'Summary'

        monthly_labels = report_data['monthly_labels']

        headers = [
            '№',
            'Студент',
            'Средний текущий балл',
            *monthly_labels,
            'Сессия 1',
            'Экзамен 1',
            'Сессия 2',
            'Экзамен 2',
            'Общий средний балл',
            'Всего оценок',
        ]

        last_col = len(headers)

        summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        title_cell = summary_sheet.cell(row=1, column=1, value='Итоговый отчёт по группе')
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        title_cell.border = border

        summary_sheet.row_dimensions[1].height = 28

        info_rows = [
            ('Группа', assignment.group.name),
            ('Предмет', assignment.subject.name),
            ('Преподаватель', assignment.teacher.full_name if assignment.teacher else '—'),
            ('Кабинет', assignment.classroom.number if assignment.classroom else '—'),
        ]

        info_start_row = 3
        for offset, (label, value) in enumerate(info_rows):
            current_row = info_start_row + offset

            label_cell = summary_sheet.cell(row=current_row, column=1, value=label)
            label_cell.fill = section_fill
            label_cell.font = header_font
            label_cell.border = border
            label_cell.alignment = left_alignment

            summary_sheet.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=current_row,
                end_column=4
            )
            value_cell = summary_sheet.cell(row=current_row, column=2, value=value)
            value_cell.font = body_font
            value_cell.border = border
            value_cell.alignment = left_alignment

            # чтобы границы были на всей объединённой области
            for col in range(2, 5):
                cell = summary_sheet.cell(row=current_row, column=col)
                cell.border = border

        header_row = 8
        for col_num, header in enumerate(headers, start=1):
            cell = summary_sheet.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        summary_sheet.row_dimensions[header_row].height = 34

        row_num = header_row + 1

        for row in report_data['rows']:
            monthly_values = [row['monthly_values'].get(month, '—') for month in report_data['monthly_numbers']]

            values = [
                row['index'],
                row['student'].full_name,
                row['current_average'],
                *monthly_values,
                row['semester_1'],
                row['exam_1'],
                row['semester_2'],
                row['exam_2'],
                row['overall_average'],
                row['grades_count'],
            ]

            monthly_start_col = 4
            monthly_end_col = monthly_start_col + len(report_data['monthly_numbers']) - 1

            semester_1_col = monthly_end_col + 1
            exam_1_col = monthly_end_col + 2
            semester_2_col = monthly_end_col + 3
            exam_2_col = monthly_end_col + 4
            overall_average_col = monthly_end_col + 5

            for col_num, value in enumerate(values, start=1):
                cell = summary_sheet.cell(row=row_num, column=col_num, value=value)

                if col_num == 2:
                    cell.font = body_font
                    cell.border = border
                    cell.alignment = left_alignment

                elif col_num == 3 or col_num == overall_average_col:
                    self.apply_average_style(
                        cell=cell,
                        value=value,
                        border=border,
                        center_alignment=center_alignment,
                        body_font=body_font
                    )

                elif monthly_start_col <= col_num <= exam_2_col:
                    self.apply_grade_style(
                        cell=cell,
                        value=value,
                        border=border,
                        center_alignment=center_alignment,
                        body_font=body_font
                    )

                else:
                    cell.font = body_font
                    cell.border = border
                    cell.alignment = center_alignment

            row_num += 1

        # ширины колонок
        summary_sheet.column_dimensions['A'].width = 8
        summary_sheet.column_dimensions['B'].width = 28
        summary_sheet.column_dimensions['C'].width = 18

        current_col = 4
        for _ in report_data['monthly_numbers']:
            summary_sheet.column_dimensions[get_column_letter(current_col)].width = 12
            current_col += 1

        summary_sheet.column_dimensions[get_column_letter(current_col)].width = 12   # Сессия 1
        summary_sheet.column_dimensions[get_column_letter(current_col + 1)].width = 12  # Экзамен 1
        summary_sheet.column_dimensions[get_column_letter(current_col + 2)].width = 12  # Сессия 2
        summary_sheet.column_dimensions[get_column_letter(current_col + 3)].width = 12  # Экзамен 2
        summary_sheet.column_dimensions[get_column_letter(current_col + 4)].width = 18  # Общий средний балл
        summary_sheet.column_dimensions[get_column_letter(current_col + 5)].width = 14  # Всего оценок

        # строки с данными
        for r in range(header_row + 1, row_num):
            summary_sheet.row_dimensions[r].height = 22
            if r % 2 == 0:
                for c in range(1, last_col + 1):
                    cell = summary_sheet.cell(row=r, column=c)
                    if cell.fill.fill_type is None:
                        cell.fill = PatternFill(fill_type='solid', fgColor='F8FAFC')

        summary_sheet.freeze_panes = f'A{header_row + 1}'
        summary_sheet.auto_filter.ref = f'A{header_row}:{get_column_letter(last_col)}{max(row_num - 1, header_row)}'
        summary_sheet.sheet_view.showGridLines = False

        details_sheet = workbook.create_sheet(title='Details')

        detail_headers = [
            'Дата',
            'День недели',
            'Студент',
            'Предмет',
            'Тип оценки',
            'Период',
            'Оценка',
            'Комментарий',
        ]

        for col_num, header in enumerate(detail_headers, start=1):
            cell = details_sheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        detail_row = 2
        weekdays = {
            0: 'Понедельник',
            1: 'Вторник',
            2: 'Среда',
            3: 'Четверг',
            4: 'Пятница',
            5: 'Суббота',
            6: 'Воскресенье',
        }

        for grade in report_data['grades']:
            values = [
                grade.date.strftime('%d.%m.%Y %H:%M'),
                weekdays.get(grade.date.weekday(), '—'),
                grade.student.full_name,
                grade.subject.name,
                grade.get_grade_type_display(),
                grade.period_label,
                grade.value,
                grade.comment or '—',
            ]

            for col_num, value in enumerate(values, start=1):
                cell = details_sheet.cell(row=detail_row, column=col_num, value=value)

                if col_num == 7:
                    self.apply_grade_style(
                        cell=cell,
                        value=value,
                        border=border,
                        center_alignment=center_alignment,
                        body_font=body_font
                    )
                else:
                    cell.font = body_font
                    cell.border = border
                    cell.alignment = left_alignment if col_num != 1 else center_alignment

            detail_row += 1

        detail_widths = {
            1: 20,
            2: 16,
            3: 28,
            4: 22,
            5: 20,
            6: 18,
            7: 10,
            8: 45,
        }

        for col_num, width in detail_widths.items():
            details_sheet.column_dimensions[get_column_letter(col_num)].width = width

        for r in range(2, detail_row):
            details_sheet.row_dimensions[r].height = 20

            if r % 2 == 0:
                for c in range(1, 9):
                    cell = details_sheet.cell(row=r, column=c)
                    if cell.fill.fill_type is None:
                        cell.fill = PatternFill(fill_type='solid', fgColor='F8FAFC')

        details_sheet.freeze_panes = 'A2'
        details_sheet.auto_filter.ref = f'A1:H{max(detail_row - 1, 1)}'
        details_sheet.sheet_view.showGridLines = False

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = (
            f'group_summary_{assignment.group.name}_{assignment.subject.name}'
            .replace(' ', '_')
            .replace('/', '_')
        ) + '.xlsx'

        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
    
class AttendanceExportExcelView(LoginRequiredMixin, View):
    WEEKDAYS_RU = {
        0: 'Понедельник',
        1: 'Вторник',
        2: 'Среда',
        3: 'Четверг',
        4: 'Пятница',
        5: 'Суббота',
        6: 'Воскресенье',
    }

    STATUS_LABELS = {
        'present': 'Присутствовал',
        'absent': 'Отсутствовал',
        'late': 'Опоздал',
        'excused': 'Уважительная причина',
    }

    def get_queryset(self):
        user = self.request.user

        queryset = Attendance.objects.select_related(
            'student',
            'assignment',
            'assignment__teacher',
            'assignment__subject',
            'assignment__group',
            'assignment__classroom',
        ).order_by('-date')

        if user.role == 'teacher':
            queryset = queryset.filter(assignment__teacher=user)
        elif user.role == 'student':
            queryset = queryset.filter(student=user)

        group_id = self.request.GET.get('group')
        if group_id and user.role != 'student':
            queryset = queryset.filter(assignment__group_id=group_id)

        subject_id = self.request.GET.get('subject')
        if subject_id:
            queryset = queryset.filter(assignment__subject_id=subject_id)

        status_value = self.request.GET.get('status')
        if status_value:
            queryset = queryset.filter(status=status_value)

        return queryset

    def get(self, request, *args, **kwargs):
        records = self.get_queryset()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Attendance'

        title_fill = PatternFill(fill_type='solid', fgColor='1E3A8A')
        header_fill = PatternFill(fill_type='solid', fgColor='DBEAFE')

        thin_side = Side(style='thin', color='D1D5DB')
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        title_font = Font(color='FFFFFF', bold=True, size=14)
        header_font = Font(bold=True, size=11)
        body_font = Font(size=10)

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        if request.user.role == 'student':
            export_title = 'Экспорт моей посещаемости'
        elif request.user.role == 'teacher':
            export_title = 'Экспорт посещаемости преподавателя'
        else:
            export_title = 'Экспорт всей посещаемости'

        headers = [
            'Дата',
            'День недели',
            'Студент',
            'Группа',
            'Предмет',
            'Преподаватель',
            'Кабинет',
            'Статус',
            'Комментарий',
        ]

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = sheet.cell(row=1, column=1, value=export_title)
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        title_cell.border = border

        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        row_num = 3

        for record in records:
            weekday_name = self.WEEKDAYS_RU.get(record.date.weekday(), '—')

            row_data = [
                record.date.strftime('%d.%m.%Y'),
                weekday_name,
                record.student.full_name,
                record.assignment.group.name if record.assignment and record.assignment.group else '—',
                record.assignment.subject.name if record.assignment and record.assignment.subject else '—',
                record.assignment.teacher.full_name if record.assignment and record.assignment.teacher else '—',
                record.assignment.classroom.number if record.assignment and record.assignment.classroom else '—',
                self.STATUS_LABELS.get(record.status, record.status),
                record.comment or '—',
            ]

            for col_num, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.font = body_font
                cell.border = border

                if col_num in [1, 2, 8]:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment

            status_cell = sheet.cell(row=row_num, column=8)
            if record.status == 'present':
                status_cell.fill = PatternFill(fill_type='solid', fgColor='DCFCE7')
                status_cell.font = Font(size=10, bold=True, color='166534')
            elif record.status == 'late':
                status_cell.fill = PatternFill(fill_type='solid', fgColor='FEF3C7')
                status_cell.font = Font(size=10, bold=True, color='92400E')
            elif record.status == 'excused':
                status_cell.fill = PatternFill(fill_type='solid', fgColor='DBEAFE')
                status_cell.font = Font(size=10, bold=True, color='1E40AF')
            elif record.status == 'absent':
                status_cell.fill = PatternFill(fill_type='solid', fgColor='FEE2E2')
                status_cell.font = Font(size=10, bold=True, color='991B1B')

            row_num += 1

        column_widths = {
            1: 14,
            2: 18,
            3: 28,
            4: 16,
            5: 24,
            6: 28,
            7: 12,
            8: 22,
            9: 40,
        }

        for col_num, width in column_widths.items():
            sheet.column_dimensions[get_column_letter(col_num)].width = width

        sheet.freeze_panes = 'A3'
        sheet.auto_filter.ref = f'A2:I{max(row_num - 1, 2)}'
        sheet.sheet_view.showGridLines = False

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f'attendance_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response
    
class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'frontend/notifications.html'
    context_object_name = 'notifications'

    def get_queryset(self):
        return Notification.objects.filter(
            user=self.request.user
        ).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = Notification.objects.filter(
            user=self.request.user,
            is_read=False
        ).count()
        return context


class NotificationMarkReadView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        notification = get_object_or_404(
            Notification,
            pk=pk,
            user=request.user
        )
        notification.is_read = True
        notification.save(update_fields=['is_read'])

        return redirect('frontend-notifications')


class NotificationMarkAllReadView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        Notification.objects.filter(
            user=request.user,
            is_read=False
        ).update(is_read=True)

        return redirect('frontend-notifications')